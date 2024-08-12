from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import *
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
import datetime
import openpyxl
from openpyxl.styles import Font
import schedule
import sys

def iniciar_driver():
    try:
        chrome_options = Options()

        arguments = [
        '--lang=pt-BR',
        '--window-size=1200,800',
        '--incognito',
        '--disable-infobars'
        '--force-device-scale-factor=0.8'
        ]

        for argument in arguments:
            chrome_options.add_argument(argument)

        chrome_options.add_experimental_option('prefs', {
            'download.prompt_for_download': False,
            'profile.default_content_setting_values.notifications': 2,
            'profile.default_content_setting_values.automatic_downloads': 1,
        })

        print("Iniciando o driver do Chrome...")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        
        wait = WebDriverWait(
            driver=driver,
            timeout=10,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException
            ]
        )

        print("Driver iniciado com sucesso.")
        return driver, wait

    except WebDriverException as e:
        print(f'Erro ao iniciar o driver: {e}')
        return None, None

def open_url(url):
    try:
        driver, wait = iniciar_driver()
        print(f"Abrindo a URL: {url}")
        driver.get(url)
        print("URL aberta com sucesso.")
        return driver, wait
    
    except Exception as e:
        print(f'Erro ao abrir a URL: {e}')
        return None, None

def get_product_name(wait, produtc_name_xpath):
    try:
        print("Buscando o nome do produto...")
        product_name = wait.until(EC.visibility_of_all_elements_located((By.XPATH, produtc_name_xpath)))
        print("Nome do produto encontrado.")
        return product_name[0]
    except TimeoutException:
        print("Erro: O nome do produto não foi encontrado no tempo esperado.")
        return None
    except Exception as e:
        print(f"Erro inesperado ao buscar o nome do produto: {e}")
        return None

def get_product_price(wait, product_price_xpath):
    try:
        print("Buscando o preço do produto...")
        product_price = wait.until(EC.visibility_of_all_elements_located((By.XPATH, product_price_xpath)))
        print("Preço do produto encontrado.")
        return product_price[0]
    except TimeoutException:
        print("Erro: O preço do produto não foi encontrado no tempo esperado.")
        return None
    except Exception as e:
       print(f"Erro inesperado ao buscar o preço do produto: {e}")
       return None 

def get_current_datetime():
    now = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
    return now

def write_to_excel(data, link_text, hyperlink, filename='monitoramento_preco.xlsx'):
    try:
        print("Escrevendo dados no arquivo Excel...")
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        print("Arquivo Excel não encontrado. Criando um novo...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Nome do produto', 'Data', 'Preço', 'Link'])

    sheet.append(data)
    hyperlink_cell = sheet.cell(row=sheet.max_row, column=sheet.max_column)
    hyperlink_cell.value = link_text
    hyperlink_cell.hyperlink = hyperlink
    hyperlink_cell.font = Font(color='0000FF', underline='single')
    workbook.save(filename)

def split_product_price(product_price_text):
    return float(product_price_text.split()[1].replace('.', '').replace(',', '.')) 

def schedule_operation(minutes):
    schedule.every(minutes).minutes.do(main)

    while True:
        schedule.run_pending()
        sleep(1)

def main():

    url = "https://www.americanas.com.br/produto/6682125344/console-playstation-5-standard-edition-branco-controle-sem-fio-dualsense-branco?pfm_index=5&pfm_page=hotsite&pfm_pos=grid&pfm_type=hotsite_page&offerId=6593d3e3cc5530938503f639&cor=Branco&cross%20docking=1&condition=NEW"
    product_price_xpath = "//div[@class='styles__PriceText-sc-1o94vuj-0 kbIkrl priceSales']"
    produtc_name_xpath = "//h1[@class='product-title__Title-sc-1hlrxcw-0 jyetLr']"
    link_text = 'Playstation 5'

    print("Iniciando monitoramento do produto...")

    driver, wait = open_url(url)
    if not driver or not wait:
        print("Erro ao inicializar o driver ou abrir a URL. Encerrando a execução.")
        sys.exit(1)

    product_name = get_product_name(wait, produtc_name_xpath)
    if not product_name:
        print("Não foi possível obter o nome do produto. Encerrando a execução.")
        driver.close()
        sys.exit(1)

    product_price = get_product_price(wait, product_price_xpath)
    if not product_price:
        print("Não foi possível obter o preço do produto. Encerrando a execução.")
        driver.close()
        sys.exit(1)

    now = get_current_datetime()
    product_price_float = split_product_price(product_price.text)
    excel_file = write_to_excel([product_name.text, now, product_price_float, url],link_text, url)
    
    print("Fechando o driver...")
    driver.close()
    print("Driver fechado. Monitoramento concluído.")

if __name__ == '__main__':
    schedule_operation(minutes=30)